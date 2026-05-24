import openpyxl

from config.config_reader import ConfigReader
from utils.logger import get_logger

logger = get_logger(__name__)


class ExcelReader:
    def __init__(self, filepath=None):
        self.filepath = filepath or ConfigReader.get_testdata_path()
        self.workbook = None
        self.sheet = None

    def load_sheet(self, sheet_name="Sheet1"):
        try:
            self.workbook = openpyxl.load_workbook(self.filepath)
            if sheet_name in self.workbook.sheetnames:
                self.sheet = self.workbook[sheet_name]
            else:
                self.sheet = self.workbook[self.workbook.sheetnames[0]]
            logger.info(f"Loaded sheet '{self.sheet.title}' from {self.filepath}")
        except Exception as error:
            logger.error(f"Failed to load sheet: {error}")
            raise

    def get_all_data(self):
        headers = [str(cell.value).strip().lower() for cell in self.sheet[1]]
        data = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                data.append(dict(zip(headers, row)))
        return data

    def get_passenger_details(self):
        self.load_sheet("PassengerDetails")
        all_data = self.get_all_data()
        if all_data:
            row = all_data[0]
            return {
                "name": row.get("name"),
                "age": row.get("age"),
                "gender": row.get("gender", "Female"),
                "email": row.get("email"),
                "phone": row.get("phone"),
            }
        raise ValueError("No passenger data found in Excel sheet")

    def close(self):
        if self.workbook:
            self.workbook.close()